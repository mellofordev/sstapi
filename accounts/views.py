from rest_framework.decorators import api_view,authentication_classes
from rest_framework.response import Response
from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token
from rest_framework.authentication import TokenAuthentication
from django.shortcuts import redirect
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Profile
from artsapi.models import Team,Program
from .serializers import ProfileSerializer

# Create your views here.
@api_view(['POST'])
def signup_view(request):
    if request.method=='POST':
        response={}
        print("activated")
        #userData=Client.login(request.data['username'],request.data['username'])
        if request.data['isLoggedIn']=='True':
            try:
                get_user_ = User.objects.get(username=request.data['username'])
                print(get_user_)
                token,obj=Token.objects.get_or_create(user=get_user_)
                return Response({"token":token.key})
            except ObjectDoesNotExist:   
                user=User(
                username=request.data['username'],
                email='dummy@sctarts.com'
                )
                password =request.data['password']
                user.set_password(password)
                user.save()
                get_user=User.objects.get(username=request.data['username'])
                print(get_user)
                get_user.profile.name = request.data['name']
                if request.data['gender']=='Male':
                    get_user.profile.gender='m'
                else:
                    get_user.profile.gender='f'
                get_user_department=request.data['department_id']
                if(len(get_user_department)==10):
                    get_user_department=request.data['department_id'][5:7]
                    if( get_user_department=="ME" or get_user_department=="EC" ):
                        get_user.profile.department='default'
                    else:
                        get_user.profile.department=request.data['department_id'][5:7]
                    get_user.profile.year=4-int(request.data['department_id'][4:5])
                elif(len(get_user_department)==11):
                    get_user_department=request.data['department_id'][6:8]
                    if( get_user_department=="ME" or get_user_department=="EC" ):
                        get_user.profile.department='default'
                    else:
                        get_user.profile.department=request.data['department_id'][5:7]
                    get_user.profile.year=4-int(request.data['department_id'][5:6])
                get_user.profile.chest_number=get_user.id
                get_user.save()
                token,obj=Token.objects.get_or_create(user=get_user)
                response['token']=str(token)
                
        else:
            return Response({"error":"Check your username or password !"})
        return Response(response)
def client_signup(request):
    return redirect('http://localhost:8501')

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
def profile_api(request):
    get_user=request.user
    if get_user.is_anonymous:
        return Response({"error":'Token not provided '})
    try:
        profile = Profile.objects.get(user=get_user)
        serializer= ProfileSerializer(profile)
    except ObjectDoesNotExist:
        return Response({"error":"Token not provided"})
    return Response({"data":serializer.data})

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
def profile_department_set_api(request,slug):
    get_user=request.user 
    if get_user.is_anonymous:
        return Response({'error':'Token not provided'})
    try:
        profile = get_user.profile
        if profile.department=='default':
            profile.department=slug
            print(slug)
            profile.save()
            return Response({'data':'updated'})
        else:
            return Response({'data':"Cannot set department for this user"})
    except ObjectDoesNotExist:
        return Response({'error':'User doest not exists'})

def export_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="am_filtered_users.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Profile"

    headers = ["lead_name", "department", "program", "members"]
    ws.append(headers)

    profiles = Profile.objects.filter(department='AM')
    for profile in profiles:
        if len(profile.registered_events.all())!=0:
            registered_events = ', '.join([event.name for event in profile.registered_events.all()])
            ws.append([profile.name, profile.department, registered_events, len(profile.registered_events.all())])

    wb.save(response)
    return response
    
def export_team_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="am_filtered_group_users.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Profile"

    headers = ["lead_name", "department", "program", "members"]
    ws.append(headers)

    teams=Team.objects.all()
    for team in teams:
        if len(team.members.all())!=0 and team.team_lead.profile.department=='AM':
            members = ', '.join([profile.name for profile in team.members.all()])
            ws.append([team.team_lead.profile.name, team.team_lead.profile.department,
                        team.program.name,members
                       
                       ])
    wb.save(response)
    return response
def export_program_data(request):
    program = Program.objects.get(name='program')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{program.name}_registered_users_sheet.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Program"

    headers = ["program_name", "registered_user_name", "chest_number","department"]  
    ws.append(headers)

    for profile in program.registered_users.all():
        ws.append([program.name, profile.name, profile.chest_number,profile.department])  

    wb.save(response)
    return response
def export_team_program_data(request):
    program = Program.objects.get(name='team')
    teams = Team.objects.all()
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{program.name}_registered_users_sheet.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Team-Events"

    headers = ["program_name", "registered_team_lead", "chest_number","department"]  
    ws.append(headers)

    for profile in program.registered_users.all():
        for team in teams:
            if team.team_lead.profile.chest_number == profile.chest_number and team.program.name==program.name:
                ws.append([program.name, profile.name, profile.chest_number,team.team_lead.profile.department])  

    wb.save(response)
    return response

def export_individual_results(request):
    programs = Program.objects.filter(program_type='s')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="individual_results.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Individual-Results"

    headers = ["Program Name", "Winner Type", "Name", "Chest Number", "Department"]
    ws.append(headers)

    for program in programs:
        for winner_type, winner_set in [('First', program.winner_first.all()), ('Second', program.winner_second.all()), ('Third', program.winner_third.all())]:
            for winner in winner_set:
                ws.append([program.name, winner_type, winner.name, winner.chest_number, winner.department])

    wb.save(response)
    return response

def export_group_results(request):
    programs = Program.objects.filter(program_type='g')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="group_results.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Group-Results"

    headers = ["Program Name", "Team Lead", "Team Members", "Department"]
    ws.append(headers)

    for program in programs:
        teams = Team.objects.filter(program=program)
        for team in teams:
            team_members = ', '.join([member.name for member in team.members.all()])
            ws.append([program.name, team.team_lead.profile.name, team_members, team.team_lead.profile.department])

    wb.save(response)
    return response